"""
company.py
----------
Full company-page scraper + report exporters for one DSE stock.

Given a trading code, this module downloads EVERYTHING shown on the
stock's official DSE company page

    https://www.dsebd.org/displayCompany.php?name=<CODE>

— market information, basic information, dividends, interim / audited
financial performance, P/E tables, other information, corporate
performance and the company address — plus the three history graphs
linked on that page (Closing Price / Total Trade / Total Volume) at the
"2 years" duration, whose data points are embedded in the DSE graph
popup pages.

Report layout
    Both exports open with the report metadata and a numbered CONTENTS
    list (so the three graph series are announced up front), then one
    numbered section per page heading. DSE's label-value tables (which
    the site prints as ``label | value | label | value`` grids) are
    unfolded into a tidy two-column ``Field / Value`` layout; genuine
    matrix tables (financial performance, P/E history) keep their
    original rows and columns. The Excel workbook adds an *Overview*
    sheet (metadata + contents), a *Company Info* sheet with the
    numbered sections, and one sheet per graph.

The bundle is exposed as ready-to-download CSV or Excel bytes for the
dashboard cards' Download button. Fetches are cached for a few minutes
so CSV + Excel clicks on the same stock reuse one scrape. The public
``csv_download_bytes`` / ``excel_download_bytes`` helpers never raise —
if dsebd.org is unreachable they return a small file explaining the
failure (they run inside Streamlit's deferred download callable, where
an exception would surface as a broken download).
"""

from __future__ import annotations

import csv
import io
import logging
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from typing import Callable, Dict, List, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup

from market import HEADERS, _ca_bundle_path
from utils import DHAKA_FMT

logger = logging.getLogger(__name__)

COMPANY_URL = "https://www.dsebd.org/displayCompany.php"
GRAPH_URL = "https://www.dsebd.org/php_graph/monthly_graph.php"

# The three graphs offered on the company page, exported at the
# "2 years" duration (the page encodes durations in months: 24 = 2 years).
GRAPH_TYPES: List[Tuple[str, str]] = [
    ("price", "Closing Price Graph"),
    ("trd", "Total Trade Graph"),
    ("vol", "Total Volume Graph"),
]
GRAPH_DURATION_MONTHS = 24
GRAPH_DURATION_LABEL = "2 Years"

REQUEST_TIMEOUT_SECONDS = 30
CACHE_TTL_SECONDS = 300.0  # CSV + Excel of one stock share a single scrape

# Graph popups embed the series as dygraphs CSV literals:
#   "Date,Price\n" + "2024-07-14,3.6\n" + ...
_GRAPH_HEADER_RE = re.compile(r'"Date,([A-Za-z ]+)\\n"')
_GRAPH_POINT_RE = re.compile(r'"(\d{4}-\d{2}-\d{2}),(-?\d+(?:\.\d+)?)\\n"')


class CompanyFetchError(RuntimeError):
    """Raised when the company page (or a required part) cannot be fetched."""


@dataclass
class Section:
    """One heading's worth of company-page data.

    ``kind`` is ``"kv"`` when the table was a label-value grid (rows are
    2-cell ``[field, value]`` pairs) or ``"matrix"`` when it is a real
    table whose first row is the column header (financials, P/E history).
    """

    title: str
    kind: str                        # "kv" | "matrix"
    rows: List[List[str]] = field(default_factory=list)


@dataclass
class GraphSeries:
    """One history graph: title, column names and (date, value) points."""

    title: str                      # e.g. "Closing Price Graph (2 Years)"
    columns: List[str]              # e.g. ["Date", "Price"]
    points: List[Tuple[str, float]] = field(default_factory=list)


@dataclass
class CompanyReport:
    """Everything scraped from one stock's DSE company page."""

    code: str
    company_name: str
    source_url: str
    fetched_at: datetime
    sections: List[Section] = field(default_factory=list)   # page order
    graphs: List[GraphSeries] = field(default_factory=list)


# ----------------------------------------------------------------------
# Fetching
# ----------------------------------------------------------------------
def _get(url: str, params: Dict[str, object]) -> str:
    """HTTP GET with one retry. Raises CompanyFetchError on failure."""
    last_error = "unknown error"
    for attempt in (1, 2):
        try:
            resp = requests.get(
                url, params=params, headers=HEADERS,
                verify=_ca_bundle_path(), timeout=REQUEST_TIMEOUT_SECONDS,
            )
            resp.raise_for_status()
            if len(resp.text) < 500:
                last_error = "response suspiciously small (possible block page)"
            else:
                return resp.text
        except requests.exceptions.RequestException as exc:
            last_error = str(exc)
        if attempt == 1:
            time.sleep(1)
    raise CompanyFetchError(f"Could not fetch {url}: {last_error}")


def _soup(html: str) -> BeautifulSoup:
    try:
        return BeautifulSoup(html, "lxml")
    except Exception:
        return BeautifulSoup(html, "html.parser")


def _clean(text: str) -> str:
    """Collapse all whitespace runs (incl. &nbsp; / newlines) to one space."""
    return " ".join(text.split())


def _table_cells(table) -> List[List[Tuple[str, int, str]]]:
    """One table's rows as (tag, colspan, text) cells; nested tables skipped."""
    rows: List[List[Tuple[str, int, str]]] = []
    for tr in table.find_all("tr"):
        if tr.find_parent("table") is not table:
            continue  # row belongs to a table nested inside a cell
        cells = []
        for c in tr.find_all(["th", "td"], recursive=False):
            try:
                span = max(1, int(c.get("colspan", 1)))
            except (TypeError, ValueError):
                span = 1
            cells.append((c.name, span, _clean(c.get_text(" ", strip=True))))
        texts = [t for _, _, t in cells]
        # Drop the graph-selector dropdowns row ("-Select Option- 1 month
        # ...") — the graphs themselves are exported as full data series.
        if any(texts) and "-Select Option-" not in " ".join(texts):
            rows.append(cells)
    return rows


def _classify_section(raw: List[List[Tuple[str, int, str]]]) -> Tuple[str, List[List[str]]]:
    """Turn one section's raw cells into ("kv", field/value rows) if every
    row is a label-value shape, else ("matrix", rows) kept verbatim.

    DSE prints label-value data three ways: ``th,td`` /  ``td,td`` single
    pairs, ``th,td,th,td`` double pairs (two fields per visual row), and
    header-only ``th,th`` cells that embed "Label: value" text. Anything
    else (financial matrices with colspans, address rows) stays a matrix;
    matrix cells expand their colspan with blanks so columns still line up.
    """
    kv: List[List[str]] = []
    for cells in raw:
        tags = [t for t, _, _ in cells]
        texts = [x for _, _, x in cells]
        if len(cells) == 2 and tags != ["th", "th"]:
            kv.append(texts)
        elif len(cells) == 4 and tags == ["th", "td", "th", "td"]:
            kv.append(texts[:2])
            kv.append(texts[2:])
        elif (cells and all(t == "th" for t in tags)
                and all(":" in x for x in texts)):
            kv.extend([p.strip() for p in x.split(":", 1)] for x in texts)
        else:
            matrix = []
            for row_cells in raw:
                row: List[str] = []
                for _, span, text in row_cells:
                    row.append(text)
                    row.extend([""] * (span - 1))
                matrix.append(row)
            return "matrix", matrix
    return "kv", [row for row in kv if any(row)]


def _parse_company_page(html: str, code: str) -> Tuple[str, List[Section]]:
    """Extract (company name, sections in page order) from the HTML.

    The company content is a sequence of ``h2.BodyHead`` headings, each
    followed by one or more ``table#company`` tables. Everything before
    the "Company Name:" heading is site chrome (ticker, sidebar) and is
    skipped; every cell after it is kept.
    """
    soup = _soup(html)
    company_name = code
    sections: List[Section] = []
    current_title: Optional[str] = None
    current_raw: List[List[Tuple[str, int, str]]] = []
    started = False

    def flush() -> None:
        nonlocal current_title, current_raw
        if current_title and current_raw:
            kind, rows = _classify_section(current_raw)
            sections.append(Section(title=current_title, kind=kind, rows=rows))
        current_raw = []

    for el in soup.find_all(["h2", "table"]):
        if el.name == "h2" and "BodyHead" in (el.get("class") or []):
            title = _clean(el.get_text(" ", strip=True))
            if title.startswith("Company Name:"):
                started = True
                company_name = title.partition(":")[2].strip() or code
                flush()
                current_title = "Company Identity"
                continue
            if started:
                flush()
                current_title = title
        elif started and el.name == "table" and el.get("id") == "company":
            current_raw.extend(_table_cells(el))
    flush()
    return company_name, sections


def _fetch_graph(code: str, graph_type: str, graph_name: str) -> GraphSeries:
    """Fetch one graph popup and pull its embedded (date, value) series."""
    title = f"{graph_name} ({GRAPH_DURATION_LABEL})"
    html = _get(GRAPH_URL, {
        "inst": code,
        "duration": GRAPH_DURATION_MONTHS,
        "type": graph_type,
    })
    header = _GRAPH_HEADER_RE.search(html)
    value_col = header.group(1).strip() if header else graph_name
    points = [(date, float(value))
              for date, value in _GRAPH_POINT_RE.findall(html)]
    return GraphSeries(title=title, columns=["Date", value_col], points=points)


def fetch_company_report(code: str) -> CompanyReport:
    """Scrape the full company page + the three 2-year graphs. May raise."""
    code = code.strip().upper()
    source_url = f"{COMPANY_URL}?name={code}"
    html = _get(COMPANY_URL, {"name": code})
    company_name, sections = _parse_company_page(html, code)
    if not sections:
        raise CompanyFetchError(
            f"No company data found on {source_url} — the page layout may "
            "have changed or the trading code is not listed."
        )

    graphs: List[GraphSeries] = []
    for graph_type, graph_name in GRAPH_TYPES:
        try:
            graphs.append(_fetch_graph(code, graph_type, graph_name))
        except CompanyFetchError as exc:
            # Keep the report useful even if one graph endpoint hiccups.
            logger.warning("Graph '%s' for %s failed: %s", graph_type, code, exc)
            graphs.append(GraphSeries(
                title=f"{graph_name} ({GRAPH_DURATION_LABEL})",
                columns=["Date", graph_name],
            ))

    return CompanyReport(
        code=code,
        company_name=company_name,
        source_url=source_url,
        fetched_at=datetime.now(ZoneInfo("Asia/Dhaka")),
        sections=sections,
        graphs=graphs,
    )


# ----------------------------------------------------------------------
# Small TTL cache — one scrape serves both the CSV and the Excel click.
# ----------------------------------------------------------------------
_cache: Dict[str, Tuple[float, CompanyReport]] = {}
_cache_lock = threading.Lock()


def get_company_report(code: str) -> CompanyReport:
    """Cached ``fetch_company_report`` (TTL a few minutes). May raise."""
    key = code.strip().upper()
    with _cache_lock:
        hit = _cache.get(key)
        if hit and time.monotonic() - hit[0] < CACHE_TTL_SECONDS:
            return hit[1]
    report = fetch_company_report(key)
    with _cache_lock:
        _cache[key] = (time.monotonic(), report)
    return report


# ----------------------------------------------------------------------
# Exporters
# ----------------------------------------------------------------------
def _meta_rows(report: CompanyReport) -> List[List[str]]:
    return [
        ["DSE COMPANY REPORT"],
        ["Trading Code", report.code],
        ["Company Name", report.company_name],
        ["Source", report.source_url],
        ["Downloaded (Asia/Dhaka)", report.fetched_at.strftime(DHAKA_FMT)],
    ]


def _contents(report: CompanyReport) -> List[Tuple[int, str]]:
    """Numbered list of every block in the report: sections, then graphs."""
    items = [(i, s.title) for i, s in enumerate(report.sections, 1)]
    offset = len(items)
    for j, g in enumerate(report.graphs, 1):
        items.append((offset + j,
                      f"{g.title} — {len(g.points)} daily data points"))
    return items


def _write_report_csv_body(writer, report: CompanyReport) -> None:
    """Write one report's CONTENTS + numbered sections + graphs to a writer.

    Shared by the single-stock export and each stock's block inside the
    bulk export, so both files carry the identical structure.
    """
    writer.writerow(["CONTENTS"])
    for number, title in _contents(report):
        writer.writerow([number, title])

    for number, section in enumerate(report.sections, 1):
        writer.writerow([])
        writer.writerow([f"SECTION {number} — {section.title}"])
        if section.kind == "kv":
            writer.writerow(["Field", "Value"])
        writer.writerows(section.rows)

    offset = len(report.sections)
    for number, graph in enumerate(report.graphs, 1):
        writer.writerow([])
        writer.writerow([f"SECTION {offset + number} — {graph.title}"])
        writer.writerow(graph.columns)
        if graph.points:
            writer.writerows(graph.points)
        else:
            writer.writerow(["No data available"])


def report_to_csv_bytes(report: CompanyReport) -> bytes:
    """One CSV: metadata, CONTENTS, then every numbered section and graph.

    Encoded UTF-8 with BOM so Excel opens it with the right charset.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf)
    writer.writerows(_meta_rows(report))
    writer.writerow([])
    _write_report_csv_body(writer, report)
    return buf.getvalue().encode("utf-8-sig")


def report_to_excel_bytes(report: CompanyReport) -> bytes:
    """Workbook: Overview (meta + contents), Company Info, one sheet/graph."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", start_color="4F46E5")   # indigo banner
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill("solid", start_color="6366F1")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", start_color="EEF2FF")

    def banner(ws, text: str, cols: int = 2, big: bool = False) -> None:
        """Append a filled title row spanning ``cols`` cells."""
        ws.append([text])
        r = ws.max_row
        for i in range(1, cols + 1):
            cell = ws.cell(row=r, column=i)
            cell.font = title_font if big else section_font
            cell.fill = title_fill if big else section_fill

    def header(ws, values: List[str]) -> None:
        """Append a bold, tinted column-header row."""
        ws.append(values)
        r = ws.max_row
        for i in range(1, len(values) + 1):
            cell = ws.cell(row=r, column=i)
            cell.font = header_font
            cell.fill = header_fill

    wb = Workbook()

    # --- Overview: metadata + contents, so the graphs are announced ---
    ws = wb.active
    ws.title = "Overview"
    banner(ws, "DSE COMPANY REPORT", cols=3, big=True)
    for row in _meta_rows(report)[1:]:
        ws.append(row)
    ws.append([])
    header(ws, ["#", "Contents", "Location"])
    n_sections = len(report.sections)
    for number, title in _contents(report):
        where = ("Company Info sheet" if number <= n_sections
                 else "its own sheet (see tabs below)")
        ws.append([number, title, where])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 30

    # --- Company Info: every page section, numbered, in page order ---
    ws = wb.create_sheet("Company Info")
    max_cols = 2
    for number, section in enumerate(report.sections, 1):
        if number > 1:
            ws.append([])
        width = (2 if section.kind == "kv"
                 else max((len(r) for r in section.rows), default=2))
        max_cols = max(max_cols, width)
        banner(ws, f"{number}. {section.title}", cols=width)
        if section.kind == "kv":
            header(ws, ["Field", "Value"])
            body = section.rows
        else:
            # A matrix's first row is its natural column header.
            header(ws, section.rows[0])
            body = section.rows[1:]
        for row in body:
            ws.append(row)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 52
    for i in range(3, max_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # --- One sheet per graph: clean Date/value series -----------------
    for graph in report.graphs:
        # Sheet names must be ≤31 chars — "Closing Price Graph (2 Years)" fits.
        sheet = wb.create_sheet(graph.title[:31])
        header(sheet, graph.columns)
        if graph.points:
            for date, value in graph.points:
                sheet.append([date, value])
        else:
            sheet.append(["No data available"])
        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 16
        sheet.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ----------------------------------------------------------------------
# Never-raising entry points for the cards' deferred download buttons.
# ----------------------------------------------------------------------
def _error_rows(code: str, exc: Exception) -> List[List[str]]:
    return [
        [f"DSE Company Report — {code.upper()}"],
        ["ERROR", "The report could not be built."],
        ["Reason", str(exc)],
        ["Hint", "dsebd.org may be slow or unreachable — please try again "
                 "in a moment."],
    ]


def csv_download_bytes(code: str) -> bytes:
    """CSV report bytes for a trading code; an error CSV if the scrape fails."""
    try:
        return report_to_csv_bytes(get_company_report(code))
    except Exception as exc:
        logger.error("CSV report for %s failed: %s", code, exc)
        buf = io.StringIO(newline="")
        csv.writer(buf).writerows(_error_rows(code, exc))
        return buf.getvalue().encode("utf-8-sig")


def excel_download_bytes(code: str) -> bytes:
    """Excel report bytes for a trading code; an error sheet if it fails."""
    try:
        return report_to_excel_bytes(get_company_report(code))
    except Exception as exc:
        logger.error("Excel report for %s failed: %s", code, exc)
        from openpyxl import Workbook

        wb = Workbook()
        for row in _error_rows(code, exc):
            wb.active.append(row)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()


# ======================================================================
# Bulk export — many stocks (up to the whole market) in ONE file.
# Used by views/bulk_download.py. Each stock's block/sheet carries the
# same structure as the single-stock export above.
# ======================================================================
# Concurrency for the bulk scrape. Each stock needs 4 requests (company
# page + 3 graphs); a small worker pool keeps a whole-market export
# feasible (~minutes) while staying gentle on dsebd.org.
BULK_MAX_WORKERS = 6

BULK_SOURCE = "https://www.dsebd.org"

# One failed stock recorded as (trading code, reason).
Failure = Tuple[str, str]


def fetch_reports_bulk(
    codes: Sequence[str],
    progress_cb: Optional[Callable[[int, int, str], None]] = None,
    max_workers: int = BULK_MAX_WORKERS,
) -> Tuple[List[CompanyReport], List[Failure]]:
    """Fetch many stocks' full company reports concurrently.

    Returns ``(reports, failures)``: reports in the order the codes were
    given, failures as ``(code, reason)`` for stocks whose scrape failed
    (the rest of the bundle still succeeds). ``progress_cb(done, total,
    code)`` is invoked after each stock completes — the bulk page uses it
    to drive a progress bar. Individual reports go through the module's
    TTL cache, so a CSV build followed by an Excel build (or a retry
    after a partial failure) reuses the scrapes it already has.
    """
    ordered = list(dict.fromkeys(
        c.strip().upper() for c in codes if c and c.strip()))
    total = len(ordered)
    results: Dict[str, CompanyReport] = {}
    errors: Dict[str, str] = {}
    done = 0
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(get_company_report, c): c for c in ordered}
        for future in as_completed(futures):
            code = futures[future]
            try:
                results[code] = future.result()
            except Exception as exc:  # keep going — report the stragglers
                logger.warning("Bulk report for %s failed: %s", code, exc)
                errors[code] = str(exc)
            done += 1
            if progress_cb:
                progress_cb(done, total, code)
    reports = [results[c] for c in ordered if c in results]
    failures = [(c, errors[c]) for c in ordered if c in errors]
    return reports, failures


def _bulk_meta_rows(reports: List[CompanyReport],
                    failures: Sequence[Failure]) -> List[List[object]]:
    rows: List[List[object]] = [
        ["DSE BULK COMPANY REPORT"],
        ["Stocks included", len(reports)],
    ]
    if failures:
        rows.append(["Stocks failed", len(failures)])
    rows += [
        ["Source", BULK_SOURCE],
        ["Downloaded (Asia/Dhaka)",
         datetime.now(ZoneInfo("Asia/Dhaka")).strftime(DHAKA_FMT)],
    ]
    return rows


def reports_to_csv_bytes(reports: List[CompanyReport],
                         failures: Sequence[Failure] = ()) -> bytes:
    """ONE CSV bundling many stocks' full company reports.

    Opens with the bundle metadata, a CONTENTS list of every stock, and
    (if any) the failed stocks with reasons. Then one block per stock —
    each block identical in structure to the single-stock CSV: source
    line, CONTENTS, numbered sections, then the three 2-year graph
    series. UTF-8 with BOM so Excel opens it with the right charset.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf)
    writer.writerows(_bulk_meta_rows(reports, failures))
    writer.writerow([])
    writer.writerow(["CONTENTS"])
    writer.writerow(["#", "Trading Code", "Company Name", "Report blocks"])
    for i, r in enumerate(reports, 1):
        writer.writerow([i, r.code, r.company_name,
                         f"{len(r.sections)} sections + {len(r.graphs)} graphs"])
    if failures:
        writer.writerow([])
        writer.writerow(["FAILED STOCKS"])
        writer.writerow(["Trading Code", "Reason"])
        for code, reason in failures:
            writer.writerow([code, reason])

    for i, r in enumerate(reports, 1):
        writer.writerow([])
        writer.writerow(["=" * 72])
        writer.writerow([f"STOCK {i} OF {len(reports)} — {r.code} · "
                         f"{r.company_name}"])
        writer.writerow(["Source", r.source_url])
        writer.writerow(["Fetched (Asia/Dhaka)",
                         r.fetched_at.strftime(DHAKA_FMT)])
        writer.writerow([])
        _write_report_csv_body(writer, r)
    return buf.getvalue().encode("utf-8-sig")


# Fixed schema of the flat (analysis-ready) bulk CSV: one row per data
# point, so the whole export is a single machine-parsable table.
FLAT_CSV_COLUMNS = ["Trading Code", "Company Name", "Section", "Item",
                    "Field", "Value"]

_NUMBER_RE = re.compile(r"-?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?")


def _coerce_number(text: str):
    """'1,234.50' → 1234.5 so columns load as numbers, not text.

    Only fully-numeric cells are converted (thousand separators dropped);
    dates, percentages and free text pass through verbatim.
    """
    if _NUMBER_RE.fullmatch(text):
        number = float(text.replace(",", ""))
        return int(number) if number.is_integer() else number
    return text


def _flat_rows(report: CompanyReport) -> List[List[object]]:
    """One report as tidy rows matching ``FLAT_CSV_COLUMNS``.

    kv sections carry the label in Field; matrix sections use their first
    row as the column header (as the Excel export does), the row's first
    cell as Item and the header cell as Field; graph points use the date
    as Item. Every row has exactly six cells.
    """
    code, name = report.code, report.company_name
    rows: List[List[object]] = []
    for section in report.sections:
        if section.kind == "kv":
            for row in section.rows:
                rows.append([code, name, section.title, "",
                             row[0] if row else "",
                             _coerce_number(row[1]) if len(row) > 1 else ""])
        elif section.rows:
            header, body = section.rows[0], section.rows[1:]
            if not body:                    # lone row — keep it as data
                header, body = [], section.rows
            for row in body:
                item = row[0] if row else ""
                cells = [(j, cell) for j, cell in enumerate(row[1:], 1)
                         if cell != ""]
                if not cells:               # text-only line (e.g. address)
                    rows.append([code, name, section.title, "", "", item])
                for j, cell in cells:
                    field_name = (header[j] if j < len(header) and header[j]
                                  else f"Column {j + 1}")
                    rows.append([code, name, section.title, item,
                                 field_name, _coerce_number(cell)])
    for graph in report.graphs:
        value_name = graph.columns[1] if len(graph.columns) > 1 else "Value"
        for date, value in graph.points:
            rows.append([code, name, graph.title, date, value_name, value])
    return rows


def reports_to_flat_csv_bytes(reports: List[CompanyReport],
                              failures: Sequence[Failure] = ()) -> bytes:
    """ONE machine-parsable CSV: a single flat table over many stocks.

    Unlike ``reports_to_csv_bytes`` (a human-readable report layout with
    stacked per-stock blocks), this is tidy "long" data: a constant
    six-column header and one row per data point, so pandas / Excel / AI
    tools can load it directly (``pd.read_csv`` works as-is). Failed
    stocks appear as rows with Section ``FETCH FAILED``. UTF-8 with BOM
    so Excel opens it with the right charset.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf)
    writer.writerow(FLAT_CSV_COLUMNS)
    for report in reports:
        writer.writerows(_flat_rows(report))
    for code, reason in failures:
        writer.writerow([code, "", "FETCH FAILED", "", "Reason", reason])
    return buf.getvalue().encode("utf-8-sig")


def _safe_sheet_name(code: str, used: set) -> str:
    """Excel-legal, unique sheet name for a trading code (≤31 chars)."""
    name = re.sub(r"[\[\]:*?/\\]", "_", code)[:31] or "STOCK"
    base, n = name, 2
    while name in used:
        name = f"{base[:28]}_{n}"
        n += 1
    used.add(name)
    return name


def reports_to_excel_bytes(reports: List[CompanyReport],
                           failures: Sequence[Failure] = ()) -> bytes:
    """ONE workbook bundling many stocks: Overview + one sheet per stock.

    The Overview sheet lists the bundle metadata and every stock with its
    sheet name (plus any failures). Each stock sheet then carries the
    full single-stock report — banner, source, the numbered page
    sections and the three 2-year graph series — i.e. the same content
    as the per-card Excel export, condensed to one sheet per stock so a
    whole-market bundle stays navigable.

    Built with openpyxl's write-only mode: a whole-market export is
    ~400 sheets / several hundred thousand rows, which streams fine but
    would not fit comfortably in a normal in-memory cell tree.
    """
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", start_color="4F46E5")   # indigo banner
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill("solid", start_color="6366F1")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", start_color="EEF2FF")

    wb = Workbook(write_only=True)

    def styled(ws, values: List[object], font, fill, pad: int = 0) -> None:
        """Append one row with every cell styled (padded to ``pad`` cells)."""
        vals = list(values) + [""] * max(0, pad - len(values))
        row = []
        for v in vals:
            cell = WriteOnlyCell(ws, value=v)
            cell.font = font
            cell.fill = fill
            row.append(cell)
        ws.append(row)

    used_names: set = set()
    sheet_names = [_safe_sheet_name(r.code, used_names) for r in reports]

    # --- Overview: metadata + every stock and where to find it --------
    ws = wb.create_sheet("Overview")
    for col, width in (("A", 26), ("B", 20), ("C", 64), ("D", 40)):
        ws.column_dimensions[col].width = width
    meta = _bulk_meta_rows(reports, failures)
    styled(ws, meta[0], title_font, title_fill, pad=4)
    for row in meta[1:]:
        ws.append(row)
    ws.append([])
    styled(ws, ["#", "Trading Code", "Company Name", "Status / Sheet"],
           header_font, header_fill)
    for i, (r, sheet) in enumerate(zip(reports, sheet_names), 1):
        ws.append([i, r.code, r.company_name, f"sheet '{sheet}'"])
    for code, reason in failures:
        ws.append(["—", code, f"FAILED — {reason}", "not included"])

    # --- One sheet per stock: full report, sections then graphs -------
    for r, sheet_name in zip(reports, sheet_names):
        ws = wb.create_sheet(sheet_name)
        # Column widths must be set before rows in write-only mode.
        max_cols = max(
            [2] + [len(row) for s in r.sections if s.kind == "matrix"
                   for row in s.rows])
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 52
        for i in range(3, max_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

        styled(ws, [f"{r.code} — {r.company_name}"], title_font, title_fill,
               pad=2)
        ws.append(["Source", r.source_url])
        ws.append(["Fetched (Asia/Dhaka)", r.fetched_at.strftime(DHAKA_FMT)])

        for number, section in enumerate(r.sections, 1):
            ws.append([])
            width = (2 if section.kind == "kv"
                     else max((len(row) for row in section.rows), default=2))
            styled(ws, [f"{number}. {section.title}"], section_font,
                   section_fill, pad=width)
            if section.kind == "kv":
                styled(ws, ["Field", "Value"], header_font, header_fill)
                body = section.rows
            else:
                # A matrix's first row is its natural column header.
                styled(ws, section.rows[0], header_font, header_fill)
                body = section.rows[1:]
            for row in body:
                ws.append(row)

        offset = len(r.sections)
        for number, graph in enumerate(r.graphs, 1):
            ws.append([])
            styled(ws, [f"{offset + number}. {graph.title}"], section_font,
                   section_fill, pad=2)
            styled(ws, graph.columns, header_font, header_fill)
            if graph.points:
                for date, value in graph.points:
                    ws.append([date, value])
            else:
                ws.append(["No data available"])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
